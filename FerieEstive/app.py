from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, Response
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func
from datetime import datetime, timedelta, date
from functools import wraps
import os
import io

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'ferie2026-secret-change-me')

# Credenziali da variabili d'ambiente
SHARED_USERNAME = os.environ.get('APP_USERNAME', 'utenteGT')
SHARED_PASSWORD = os.environ.get('APP_PASSWORD', 'FFG_GT')
ADMIN_PASSWORD  = os.environ.get('ADMIN_PASSWORD', 'TACCHELLA2026')

# Settimane estive 2026
# Sett. 23-32: 1 giugno – 9 agosto 2026  (prima di ferragosto)
# Sett. 33   : 10-16 agosto 2026 → SETTIMANA DI FERRAGOSTO FISSA (non selezionabile)
# Sett. 34-38: 17 agosto – 20 settembre 2026  (dopo ferragosto)
SETTIMANE_PRIMA  = list(range(23, 33))   # 10 settimane (sett. 23–32)
SETTIMANE_DOPO   = list(range(34, 39))   #  5 settimane (sett. 34–38)
SETTIMANE_ESTIVE = SETTIMANE_PRIMA + SETTIMANE_DOPO  # sett. 33 esclusa (fissa)

# Ferragosto 2026 = 15 agosto (sabato) → settimana ISO 33 (10-16 ago) — FISSA PER TUTTI
# Settimana contigua PRIMA  = sett. 32 (3–9 agosto)   → selezionabile come ferragosto
# Settimana contigua DOPO   = sett. 34 (17–23 agosto) → selezionabile come ferragosto
SETTIMANA_FERRAGOSTO_FISSA = 33   # fissa per tutti, non selezionabile
SETTIMANA_PRE_FERRAGOSTO   = 32
SETTIMANA_POST_FERRAGOSTO  = 34
SETTIMANE_FERRAGOSTO       = [SETTIMANA_PRE_FERRAGOSTO, SETTIMANA_POST_FERRAGOSTO]

# Database
data_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
os.makedirs(data_dir, exist_ok=True)
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{os.path.join(data_dir, "ferie.db")}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# ─── MESI ITALIANI ────────────────────────────────────────────────────────────
MESI_IT = {
    1: 'Gennaio', 2: 'Febbraio', 3: 'Marzo',    4: 'Aprile',
    5: 'Maggio',  6: 'Giugno',   7: 'Luglio',   8: 'Agosto',
    9: 'Settembre', 10: 'Ottobre', 11: 'Novembre', 12: 'Dicembre'
}

@app.template_filter('mese_it')
def mese_it_filter(n):
    return MESI_IT.get(n, str(n))

# ─── MODELLI ──────────────────────────────────────────────────────────────────

class Dipendente(db.Model):
    __tablename__ = 'dipendenti'
    id       = db.Column(db.Integer, primary_key=True)
    nome     = db.Column(db.String(100), nullable=False)
    cognome  = db.Column(db.String(100), nullable=False)
    attivo   = db.Column(db.Boolean, default=True)

    @property
    def nome_completo(self):
        return f"{self.cognome} {self.nome}"


class Impostazioni(db.Model):
    __tablename__ = 'impostazioni'
    id                  = db.Column(db.Integer, primary_key=True)
    max_per_settimana   = db.Column(db.Integer, default=5)
    abilita_quarta      = db.Column(db.Boolean, default=True)
    note                = db.Column(db.Text, default='')


class SettimanaConfig(db.Model):
    __tablename__ = 'settimane_config'
    id               = db.Column(db.Integer, primary_key=True)
    anno             = db.Column(db.Integer, nullable=False, default=2026)
    numero_settimana = db.Column(db.Integer, nullable=False)
    disponibile      = db.Column(db.Boolean, default=True)
    __table_args__   = (db.UniqueConstraint('anno', 'numero_settimana'),)


class Scelta(db.Model):
    """
    Ogni dipendente può fare fino a 4 scelte:
    - settimana_ferragosto : obbligatoria (prima o dopo il 15 agosto)
    - settimana_aggiuntiva : facoltativa
    - settimana_riserva    : riserva per l'aggiuntiva (facoltativa)
    - settimana_quarta     : quarta settimana se c'è capienza (facoltativa)
    """
    __tablename__ = 'scelte'
    id                   = db.Column(db.Integer, primary_key=True)
    dipendente_id        = db.Column(db.Integer, db.ForeignKey('dipendenti.id'), nullable=False, unique=True)
    settimana_ferragosto = db.Column(db.Integer, nullable=True)
    settimana_aggiuntiva = db.Column(db.Integer, nullable=True)
    settimana_riserva    = db.Column(db.Integer, nullable=True)
    settimana_quarta     = db.Column(db.Integer, nullable=True)
    ts_creazione         = db.Column(db.DateTime, default=datetime.utcnow)
    ts_aggiornamento     = db.Column(db.DateTime, default=datetime.utcnow)
    dipendente           = db.relationship('Dipendente', backref=db.backref('scelta', uselist=False))

# ─── HELPER SETTIMANE ─────────────────────────────────────────────────────────

def get_week_info(num, year=2026):
    start = date.fromisocalendar(year, num, 1)
    end   = start + timedelta(days=6)
    return {
        'numero':    num,
        'inizio':    start,
        'fine':      end,
        'label':     f"Sett. {num} – {start.strftime('%d/%m')} › {end.strftime('%d/%m')}",
        'periodo':   f"{start.strftime('%d/%m')}–{end.strftime('%d/%m/%Y')}",
        'categoria': 'prima' if num in SETTIMANE_PRIMA else 'dopo',
    }

def get_estive_weeks():
    """Restituisce le settimane estive (23-38) con flag di disponibilità."""
    configs = {s.numero_settimana: s.disponibile
               for s in SettimanaConfig.query.filter_by(anno=2026).all()}
    return [
        {**get_week_info(n), 'disponibile': configs.get(n, n in SETTIMANE_ESTIVE)}
        for n in SETTIMANE_ESTIVE
    ]

def get_week_counts():
    """Conteggio totale selezioni per settimana (tutti e 4 i tipi)."""
    counts = {}
    for s in Scelta.query.all():
        for n in [s.settimana_ferragosto, s.settimana_aggiuntiva,
                  s.settimana_riserva, s.settimana_quarta]:
            if n:
                counts[n] = counts.get(n, 0) + 1
    return counts

def get_week_counts_by_type():
    """Conteggio per settimana suddiviso per tipo di scelta."""
    result = {n: {'ferragosto': 0, 'aggiuntiva': 0, 'riserva': 0, 'quarta': 0, 'totale': 0}
              for n in SETTIMANE_ESTIVE}
    for s in Scelta.query.all():
        for tipo, val in [('ferragosto', s.settimana_ferragosto),
                          ('aggiuntiva', s.settimana_aggiuntiva),
                          ('riserva',    s.settimana_riserva),
                          ('quarta',     s.settimana_quarta)]:
            if val and val in result:
                result[val][tipo]   += 1
                result[val]['totale'] += 1
    return result

# ─── DECORATORI ───────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def wrapper(*a, **kw):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*a, **kw)
    return wrapper

def employee_required(f):
    @wraps(f)
    def wrapper(*a, **kw):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        if not session.get('dipendente_id'):
            return redirect(url_for('seleziona_dipendente'))
        return f(*a, **kw)
    return wrapper

def admin_required(f):
    @wraps(f)
    def wrapper(*a, **kw):
        if not session.get('is_admin'):
            flash('Accesso riservato agli amministratori.', 'danger')
            return redirect(url_for('login'))
        return f(*a, **kw)
    return wrapper

# ─── ROUTE PRINCIPALI ─────────────────────────────────────────────────────────

@app.route('/')
def index():
    if session.get('is_admin'):
        return redirect(url_for('admin'))
    if session.get('logged_in') and session.get('dipendente_id'):
        return redirect(url_for('dashboard'))
    if session.get('logged_in'):
        return redirect(url_for('seleziona_dipendente'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        u = request.form.get('username', '').strip()
        p = request.form.get('password', '').strip()
        if u.lower() == 'admin' and p == ADMIN_PASSWORD:
            session.clear()
            session['logged_in'] = True
            session['is_admin']  = True
            flash('Accesso amministratore effettuato.', 'success')
            return redirect(url_for('admin'))
        elif u == SHARED_USERNAME and p == SHARED_PASSWORD:
            session.clear()
            session['logged_in'] = True
            session['is_admin']  = False
            flash('Accesso effettuato. Inserisci il tuo nome per continuare.', 'success')
            return redirect(url_for('seleziona_dipendente'))
        else:
            flash('Credenziali non valide. Riprova.', 'danger')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    flash('Disconnessione effettuata.', 'info')
    return redirect(url_for('login'))


@app.route('/seleziona', methods=['GET', 'POST'])
@login_required
def seleziona_dipendente():
    if request.method == 'POST':
        nome    = request.form.get('nome', '').strip()
        cognome = request.form.get('cognome', '').strip()
        if not nome or not cognome:
            flash('Inserisci nome e cognome per continuare.', 'danger')
            return render_template('selezione_dipendente.html')

        # Cerca dipendente esistente (case-insensitive)
        d = Dipendente.query.filter(
            func.lower(Dipendente.nome)    == nome.lower(),
            func.lower(Dipendente.cognome) == cognome.lower()
        ).first()

        if d and not d.attivo:
            flash("Il tuo accesso è stato disabilitato. Contatta l'amministratore.", 'danger')
            return render_template('selezione_dipendente.html')

        if not d:
            # Prima registrazione: crea il dipendente
            d = Dipendente(nome=nome.title(), cognome=cognome.title())
            db.session.add(d)
            db.session.commit()

        session['dipendente_id']   = d.id
        session['dipendente_nome'] = d.nome_completo
        return redirect(url_for('dashboard'))

    return render_template('selezione_dipendente.html')


@app.route('/dashboard')
@employee_required
def dashboard():
    dipendente   = Dipendente.query.get(session['dipendente_id'])
    impostazioni = Impostazioni.query.first()
    scelta       = Scelta.query.filter_by(dipendente_id=dipendente.id).first()
    week_counts  = get_week_counts()
    max_pw       = impostazioni.max_per_settimana if impostazioni else 5

    all_weeks = get_estive_weeks()
    for w in all_weeks:
        n = w['numero']
        w['prenotazioni'] = week_counts.get(n, 0)
        w['max']          = max_pw
        w['piena']        = w['disponibile'] and w['prenotazioni'] >= max_pw
        w['perc']         = min(100, int(w['prenotazioni'] / max_pw * 100)) if max_pw > 0 else 0

    # Settimane per il picker ferragosto (solo sett. 32 e 34)
    all_weeks_map    = {w['numero']: w for w in all_weeks}
    ferragosto_weeks = [all_weeks_map[n] for n in SETTIMANE_FERRAGOSTO if n in all_weeks_map]

    # Settimane per il calendario aggiuntiva/riserva/quarta (tutte le estive disponibili)
    settimane_prima = [w for w in all_weeks if w['categoria'] == 'prima' and w['disponibile']]
    settimane_dopo  = [w for w in all_weeks if w['categoria'] == 'dopo'  and w['disponibile']]

    # La quarta è abilitata se l'admin lo consente e ci sono settimane con posti liberi
    libere = [w for w in all_weeks if w['disponibile'] and not w['piena']]
    quarta_disponibile = bool(impostazioni and impostazioni.abilita_quarta and len(libere) > 0)

    return render_template('dashboard.html',
        dipendente=dipendente,
        ferragosto_weeks=ferragosto_weeks,
        settimane_prima=settimane_prima,
        settimane_dopo=settimane_dopo,
        scelta=scelta,
        quarta_disponibile=quarta_disponibile,
        impostazioni=impostazioni,
        settimana_pre_ferragosto=SETTIMANA_PRE_FERRAGOSTO,
        settimana_post_ferragosto=SETTIMANA_POST_FERRAGOSTO,
    )


@app.route('/salva', methods=['POST'])
@employee_required
def salva_scelta():
    dipendente   = Dipendente.query.get(session['dipendente_id'])
    impostazioni = Impostazioni.query.first()
    max_pw = impostazioni.max_per_settimana if impostazioni else 5

    ferragosto = request.form.get('settimana_ferragosto', type=int)
    aggiuntiva = request.form.get('settimana_aggiuntiva', type=int) or None
    riserva    = request.form.get('settimana_riserva',    type=int) or None
    quarta     = request.form.get('settimana_quarta',     type=int) or None

    if not ferragosto:
        flash('Devi selezionare la settimana di ferragosto.', 'danger')
        return redirect(url_for('dashboard'))

    # La settimana di ferragosto può essere SOLO sett. 32 (prima) o sett. 34 (dopo)
    if ferragosto not in SETTIMANE_FERRAGOSTO:
        flash(f'La settimana di ferragosto deve essere la sett. {SETTIMANA_PRE_FERRAGOSTO} '
              f'(prima, 3–9 ago) oppure la sett. {SETTIMANA_POST_FERRAGOSTO} (dopo, 17–23 ago).', 'danger')
        return redirect(url_for('dashboard'))

    # Unicità tra tutte le scelte
    selezionate = [s for s in [ferragosto, aggiuntiva, riserva, quarta] if s]
    if len(selezionate) != len(set(selezionate)):
        flash('Le settimane selezionate devono essere diverse tra loro.', 'danger')
        return redirect(url_for('dashboard'))

    # Verifica disponibilità
    configs = {s.numero_settimana: s.disponibile
               for s in SettimanaConfig.query.filter_by(anno=2026).all()}
    week_counts = get_week_counts()

    # Sottrai il contributo attuale del dipendente per non contare doppio
    corrente = Scelta.query.filter_by(dipendente_id=dipendente.id).first()
    if corrente:
        for n in [corrente.settimana_ferragosto, corrente.settimana_aggiuntiva,
                  corrente.settimana_riserva,    corrente.settimana_quarta]:
            if n and n in week_counts:
                week_counts[n] = max(0, week_counts[n] - 1)

    for n in selezionate:
        if not configs.get(n, False):
            flash(f'La settimana {n} non è disponibile per la selezione.', 'danger')
            return redirect(url_for('dashboard'))
        if week_counts.get(n, 0) >= max_pw:
            flash(f'La settimana {n} ha raggiunto il numero massimo di prenotazioni ({max_pw}).', 'danger')
            return redirect(url_for('dashboard'))

    if corrente:
        corrente.settimana_ferragosto = ferragosto
        corrente.settimana_aggiuntiva = aggiuntiva
        corrente.settimana_riserva    = riserva
        corrente.settimana_quarta     = quarta
        corrente.ts_aggiornamento     = datetime.utcnow()
    else:
        db.session.add(Scelta(
            dipendente_id=dipendente.id,
            settimana_ferragosto=ferragosto,
            settimana_aggiuntiva=aggiuntiva,
            settimana_riserva=riserva,
            settimana_quarta=quarta,
        ))

    db.session.commit()
    flash('Scelta salvata con successo!', 'success')
    return redirect(url_for('dashboard'))

# ─── ROUTE ADMIN ──────────────────────────────────────────────────────────────

@app.route('/admin')
@admin_required
def admin():
    dipendenti   = Dipendente.query.order_by(Dipendente.cognome, Dipendente.nome).all()
    scelte       = Scelta.query.all()
    impostazioni = Impostazioni.query.first()
    max_pw       = impostazioni.max_per_settimana if impostazioni else 5

    week_counts_type = get_week_counts_by_type()
    all_weeks = get_estive_weeks()
    for w in all_weeks:
        n = w['numero']
        ct = week_counts_type.get(n, {})
        w['ct']    = ct
        w['max']   = max_pw
        w['perc']  = min(100, int(ct.get('totale', 0) / max_pw * 100)) if max_pw > 0 else 0
        w['piena'] = ct.get('totale', 0) >= max_pw

    scelte_map    = {s.dipendente_id: s for s in scelte}
    totale_attivi = len([d for d in dipendenti if d.attivo])
    completati    = len([s for s in scelte if s.settimana_ferragosto])

    # Nomi per visualizzazione grafica (per settimana, per tipo)
    nomi_per_settimana = {n: {'ferragosto': [], 'aggiuntiva': [], 'riserva': [], 'quarta': []}
                          for n in SETTIMANE_ESTIVE}
    for s in scelte:
        nome = s.dipendente.nome_completo
        for tipo, val in [('ferragosto', s.settimana_ferragosto),
                          ('aggiuntiva', s.settimana_aggiuntiva),
                          ('riserva',    s.settimana_riserva),
                          ('quarta',     s.settimana_quarta)]:
            if val and val in nomi_per_settimana:
                nomi_per_settimana[val][tipo].append(nome)

    return render_template('admin.html',
        dipendenti=dipendenti,
        scelte_map=scelte_map,
        settimane=all_weeks,
        nomi_per_settimana=nomi_per_settimana,
        impostazioni=impostazioni,
        totale_attivi=totale_attivi,
        completati=completati,
        max_pw=max_pw,
        settimane_prima=SETTIMANE_PRIMA,
        settimane_dopo=SETTIMANE_DOPO,
    )


@app.route('/admin/dipendente', methods=['POST'])
@admin_required
def admin_dipendente():
    action = request.form.get('action')
    if action == 'toggle':
        d = Dipendente.query.get(request.form.get('id', type=int))
        if d:
            d.attivo = not d.attivo
            db.session.commit()
            flash(f'{d.nome_completo} {"attivato" if d.attivo else "disattivato"}.', 'info')
    elif action == 'delete_scelta':
        s = Scelta.query.filter_by(
            dipendente_id=request.form.get('dipendente_id', type=int)
        ).first()
        if s:
            db.session.delete(s)
            db.session.commit()
            flash('Scelta eliminata.', 'info')
    return redirect(url_for('admin'))


@app.route('/admin/impostazioni', methods=['POST'])
@admin_required
def admin_impostazioni():
    imp = Impostazioni.query.first()
    if not imp:
        imp = Impostazioni()
        db.session.add(imp)
    imp.max_per_settimana = request.form.get('max_per_settimana', type=int, default=5)
    imp.abilita_quarta    = 'abilita_quarta' in request.form
    imp.note              = request.form.get('note', '')
    db.session.commit()
    flash('Impostazioni salvate.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/settimana', methods=['POST'])
@admin_required
def admin_settimana():
    numero      = request.form.get('numero', type=int)
    disponibile = request.form.get('disponibile') == '1'
    cfg = SettimanaConfig.query.filter_by(anno=2026, numero_settimana=numero).first()
    if cfg:
        cfg.disponibile = disponibile
    else:
        db.session.add(SettimanaConfig(anno=2026, numero_settimana=numero, disponibile=disponibile))
    db.session.commit()
    return jsonify({'ok': True, 'numero': numero, 'disponibile': disponibile})


@app.route('/admin/export')
@admin_required
def admin_export():
    """Esporta le scelte in un file Excel (.xlsx) con 2 fogli."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash('openpyxl non disponibile. Installa con: pip install openpyxl', 'danger')
        return redirect(url_for('admin'))

    wb = openpyxl.Workbook()

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill    = PatternFill("solid", fgColor="EBF3FB")
    thin_side   = Side(style='thin', color='CCCCCC')
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_al   = Alignment(horizontal='center', vertical='center')

    # ── Foglio 1: Scelte dipendenti ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Scelte Dipendenti"

    headers1 = ['Cognome', 'Nome',
                'Sett. Ferragosto', 'Periodo Ferragosto',
                'Sett. Aggiuntiva', 'Periodo Aggiuntiva',
                'Riserva Aggiuntiva', 'Periodo Riserva',
                '4ª Settimana', 'Periodo 4ª',
                'Salvata il']
    ws1.append(headers1)
    for col in range(1, len(headers1) + 1):
        cell = ws1.cell(row=1, column=col)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.border    = cell_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws1.row_dimensions[1].height = 30

    def fmt(n):
        if not n:
            return ('', '')
        w = get_week_info(n)
        return str(n), f"{w['inizio'].strftime('%d/%m')}-{w['fine'].strftime('%d/%m/%Y')}"

    dipendenti = (Dipendente.query
                  .filter_by(attivo=True)
                  .order_by(Dipendente.cognome, Dipendente.nome)
                  .all())
    scelte_map = {s.dipendente_id: s for s in Scelta.query.all()}

    for ri, d in enumerate(dipendenti, 2):
        s = scelte_map.get(d.id)
        if s:
            w_f, p_f = fmt(s.settimana_ferragosto)
            w_a, p_a = fmt(s.settimana_aggiuntiva)
            w_r, p_r = fmt(s.settimana_riserva)
            w_q, p_q = fmt(s.settimana_quarta)
            salvata  = s.ts_aggiornamento.strftime('%d/%m/%Y %H:%M') if s.ts_aggiornamento else ''
        else:
            w_f = p_f = w_a = p_a = w_r = p_r = w_q = p_q = salvata = ''
        ws1.append([d.cognome, d.nome, w_f, p_f, w_a, p_a, w_r, p_r, w_q, p_q, salvata])
        fill = alt_fill if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col in range(1, len(headers1) + 1):
            cell = ws1.cell(row=ri, column=col)
            cell.fill   = fill
            cell.border = cell_border
            cell.alignment = center_al

    for i, w in enumerate([16, 14, 12, 20, 12, 20, 16, 20, 10, 20, 18], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Foglio 2: Riepilogo settimane ────────────────────────────────────────
    ws2 = wb.create_sheet("Riepilogo Settimane")
    headers2 = ['Settimana', 'Periodo', 'Categoria',
                'Ferragosto', 'Aggiuntiva', 'Riserva', '4ª Sett.', 'Totale']
    ws2.append(headers2)
    for col in range(1, len(headers2) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.border    = cell_border
        cell.alignment = center_al
    ws2.row_dimensions[1].height = 25

    counts = get_week_counts_by_type()
    fill_prima = PatternFill("solid", fgColor="FFF3E0")
    fill_dopo  = PatternFill("solid", fgColor="E3F2FD")

    for ri, n in enumerate(SETTIMANE_ESTIVE, 2):
        wi  = get_week_info(n)
        ct  = counts.get(n, {})
        cat = 'Prima di Ferragosto' if n in SETTIMANE_PRIMA else 'Dopo Ferragosto'
        ws2.append([n,
                    f"{wi['inizio'].strftime('%d/%m')}-{wi['fine'].strftime('%d/%m/%Y')}",
                    cat,
                    ct.get('ferragosto', 0), ct.get('aggiuntiva', 0),
                    ct.get('riserva', 0),    ct.get('quarta', 0),
                    ct.get('totale', 0)])
        fill = fill_prima if n in SETTIMANE_PRIMA else fill_dopo
        for col in range(1, len(headers2) + 1):
            cell = ws2.cell(row=ri, column=col)
            cell.fill   = fill
            cell.border = cell_border
            cell.alignment = center_al

    for i, w in enumerate([10, 22, 22, 12, 12, 10, 10, 10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"ferie_2026_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )

# ─── INIT DB ──────────────────────────────────────────────────────────────────

def init_db():
    db.create_all()
    if not Impostazioni.query.first():
        db.session.add(Impostazioni(max_per_settimana=5, abilita_quarta=True))
    # Abilita di default solo le settimane estive (23-38)
    existing = {s.numero_settimana
                for s in SettimanaConfig.query.filter_by(anno=2026).all()}
    for n in range(1, 53):
        if n not in existing:
            db.session.add(SettimanaConfig(
                anno=2026,
                numero_settimana=n,
                disponibile=(n in SETTIMANE_ESTIVE)
            ))
    db.session.commit()

with app.app_context():
    init_db()

# Rendi get_week_info disponibile nei template Jinja2
app.jinja_env.globals['get_week_info'] = get_week_info

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
